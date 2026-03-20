# -*- coding: utf-8 -*-
"""
HTTP 서버 모듈 (Python/Flask)
- 정적 파일 서빙
- 동행복권 최신 당첨번호: 내부 API(selectPstLt645Info.do) 단일 호출 → /api/lotto-latest
"""
import sys
import os
from pathlib import Path

# 로컬 라이브러리 경로(Lib/site-packages)가 있으면 우선 사용
BASE_DIR = Path(__file__).resolve().parent
local_lib = BASE_DIR / 'Lib' / 'site-packages'
if local_lib.exists():
    sys.path.insert(0, str(local_lib))

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except (AttributeError, OSError):
    pass

import json
import time
import datetime

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from flask import Flask, send_from_directory, jsonify, request
from utils.logger import get_logger

# --- 설정 ---
PORT = int(os.environ.get('PORT', 8000))
CORS_HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type'
}
CORS_OPTIONS_HEADERS = {**CORS_HEADERS, 'Content-Length': '0'}
SERVER_START_TIME = (datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=9)))).strftime('%Y-%m-%d %H:%M:%S')

app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')
app.config['JSON_AS_ASCII'] = False
logger = get_logger(__name__)

# Lotto023.xlsx 표준 컬럼 (조회시작/조회종료 없음, 게임선택 다음에 선택합계)
LOTTO023_CANONICAL_HEADERS = (
    '회차', '세트', '게임', '홀짝', '연속', '핫콜', '게임선택', '선택합계',
    '선택1', '선택2', '선택3', '선택4', '선택5', '선택6',
    'Perfect순위',
)


def _lotto023_pick_valid(g):
    """게임 dict의 선택1~6이 1~45 정수 6개이고 서로 중복이 없으면 True."""
    if not isinstance(g, dict):
        return False
    nums = []
    for j in range(1, 7):
        raw = g.get('선택%d' % j)
        if raw is None or str(raw).strip() == '':
            return False
        try:
            n = int(str(raw).strip())
        except (ValueError, TypeError):
            return False
        if n < 1 or n > 45:
            return False
        nums.append(n)
    return len(nums) == 6 and len(set(nums)) == 6


def _lotto023_row_dict(headers, row):
    d = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        key = str(h).strip()
        if not key:
            continue
        d[key] = row[i] if i < len(row) else None
    return d


def _migrate_lotto023_rows(headers, data_rows):
    """구 형식 시트를 표준 헤더로 맞추고, 선택합계가 비어 있으면 선택1~6으로 계산."""
    headers = [str(h).strip() if h is not None else '' for h in (headers or [])]
    canon = list(LOTTO023_CANONICAL_HEADERS)
    new_rows = []
    for r in data_rows:
        if r is None:
            continue
        r = list(r)
        rowd = _lotto023_row_dict(headers, r)
        new_r = []
        for h in canon:
            if h == '선택합계':
                v = rowd.get('선택합계')
                if v is not None and str(v).strip() != '':
                    new_r.append(str(v).strip())
                else:
                    try:
                        s = 0
                        for i in range(1, 7):
                            x = rowd.get('선택%d' % i)
                            if x is not None and str(x).strip() != '':
                                s += int(str(x).strip())
                        new_r.append(str(s) if s else '')
                    except (ValueError, TypeError):
                        new_r.append('')
            else:
                v = rowd.get(h)
                new_r.append('' if v is None else v)
        new_rows.append(new_r)
    return canon, new_rows



def _fmt_amount(val):
    """숫자를 천 단위 콤마 문자열로, 없으면 (없음)."""
    if val is None:
        return '(없음)'
    try:
        return format(int(val), ',')
    except (TypeError, ValueError):
        return '(없음)'


def _to_latest_response(parsed):
    """API 원본 데이터를 클라이언트용 규격(standard format)으로 변환. 추첨일 YYYY-MM-DD, 당첨금 천 단위 콤마."""
    d = parsed.get('drwNoDate')
    if not d:
        d = datetime.date.today().isoformat()
    amt = parsed.get('rnk1WnAmt')
    winnerAmtFmt = _fmt_amount(amt)

    out = {
        'returnValue': 'success',
        'drwNo': parsed.get('drwNo'),
        'drwNoDate': d,
        'drwtNo1': parsed['main'][0], 'drwtNo2': parsed['main'][1],
        'drwtNo3': parsed['main'][2], 'drwtNo4': parsed['main'][3],
        'drwtNo5': parsed['main'][4], 'drwtNo6': parsed['main'][5],
        'bnusNo': parsed['bnusNo'],
        'winnerAmt': winnerAmtFmt,
        'source': parsed.get('source', '동행복권 내부 API'),
    }
    # firstWinamnt, firstPrzwnerCo, firstAccumamnt, totSellamnt 추가
    out['firstWinamnt'] = parsed.get('firstWinamnt')
    out['firstWinamntFmt'] = _fmt_amount(parsed.get('firstWinamnt'))
    out['firstPrzwnerCo'] = parsed.get('firstPrzwnerCo')
    out['firstPrzwnerCoFmt'] = str(parsed.get('firstPrzwnerCo')) if parsed.get('firstPrzwnerCo') is not None else '(없음)'
    out['firstAccumamnt'] = parsed.get('firstAccumamnt')
    out['firstAccumamntFmt'] = _fmt_amount(parsed.get('firstAccumamnt'))
    out['totSellamnt'] = parsed.get('totSellamnt')
    out['totSellamntFmt'] = _fmt_amount(parsed.get('totSellamnt'))
    return out


def fetch_latest_lotto(force=False):
    """
    동행복권 당첨번호 조회.
    1. force=True 이거나, 로컬이 최신이 아니면 API 호출.
    2. 로컬 파일(Lotto645.json)에서 최신 회차 확인.
    3. 오늘(KST)이 추첨일(토요일)이고 20:45 KST 이후면 항상 API 호출(동행복권 반영 시점).
    4. 그 외엔 기존 로직대로 필요 시에만 API 호출.
    """
    import datetime

    # KST 기준 현재 시각 (서버 타임존과 무관하게 추첨일 판단)
    kst = datetime.timezone(datetime.timedelta(hours=9))
    now_kst = datetime.datetime.now(kst)
    today_kst = now_kst.date()

    # 1. 로컬 데이터 확인
    local_parsed = None
    try:
        json_path = (BASE_DIR / '.source' / 'Lotto645.json').resolve()
        if json_path.is_file():
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data and len(data) > 0:
                    latest_local = data[0] if data[0].get('회차') else None
                    if latest_local:
                        local_parsed = {
                            'drwNo': int(latest_local.get('회차')),
                            'drwNoDate': latest_local.get('날짜'),
                            'main': [int(latest_local.get(f'번호{i}')) for i in range(1, 7)],
                            'bnusNo': int(latest_local.get('보너스번호')),
                            'rnk1WnAmt': latest_local.get('1등 당첨금액'),
                            'firstWinamnt': latest_local.get('1등 당첨금액'),
                            'firstPrzwnerCo': latest_local.get('1등 당첨자 수'),
                            'firstAccumamnt': latest_local.get('1등 총당첨금액'),
                            'totSellamnt': latest_local.get('전체 당첨상금 총액'),
                            'source': 'Lotto645.json (로컬)',
                        }
    except Exception as e:
        logger.warning('[Lotto645] 로컬 JSON 확인 중 오류: %s', e)

    # 2. API 호출 필요 여부 판단
    need_api_call = force or (local_parsed is None)
    if not need_api_call and local_parsed:
        try:
            last_date_str = local_parsed.get('drwNoDate')
            if last_date_str:
                last_date = datetime.datetime.strptime(last_date_str.strip()[:10], '%Y-%m-%d').date()
                # 마지막 회차 날짜가 오늘(KST)보다 미래면 API 불필요
                if last_date >= today_kst:
                    need_api_call = False
                else:
                    # 다음 추첨일(토요일) 계산
                    days_diff = (5 - last_date.weekday()) % 7
                    if days_diff == 0:
                        days_diff = 7
                    next_draw_date = last_date + datetime.timedelta(days=days_diff)

                    if today_kst < next_draw_date:
                        need_api_call = False
                    elif today_kst == next_draw_date:
                        # 오늘이 추첨일(토요일). 20:45 KST 이후면 항상 API 호출
                        if now_kst.hour > 20 or (now_kst.hour == 20 and now_kst.minute >= 45):
                            need_api_call = True
                        else:
                            need_api_call = False
                    else:
                        need_api_call = True
            else:
                need_api_call = True
        except Exception as e:
            logger.warning('[Lotto645] 날짜 계산 오류, API 호출 시도: %s', e)
            need_api_call = True

    if not need_api_call and local_parsed:
        logger.info('[동행복권] 로컬 데이터가 최신입니다. API 호출 생략.')
        return _to_latest_response(local_parsed), None

    # 3. API 호출 (로컬 데이터가 없거나, 갱신이 필요한 경우)
    try:
        from utils.get_lotto_round import get_latest_lotto
        parsed = get_latest_lotto()
        if parsed:
            logger.info('[동행복권] API 조회 성공: %s회 (%s)', parsed.get('drwNo'), parsed.get('drwNoDate'))
            # 로컬 파일보다 더 최신이면 자동 저장 로직 실행
            if local_parsed:
                 api_drw = parsed.get('drwNo')
                 local_drw = local_parsed.get('drwNo')
                 if api_drw and local_drw and api_drw > local_drw:
                      logger.info('[동행복권] 새 회차 발견! 로컬 파일 업데이트 시도...')
                      # 이미 가져온 데이터를 넘겨서 중복 호출 방지!
                      _update_latest_lotto645(parsed)

            return _to_latest_response(parsed), None
        
        # API 실패 시, 로컬 데이터라도 있으면 반환
        if local_parsed:
            logger.warning('[동행복권] API 실패하여 기존 로컬 데이터 반환.')
            return _to_latest_response(local_parsed), None

        return None, '동행복권 API에서 당첨 정보를 가져오지 못했습니다.'
    except Exception as e:
        if local_parsed:
             logger.warning('동행복권 API 호출 중 예외 발생. 로컬 데이터 반환: %s', e)
             return _to_latest_response(local_parsed), None
        logger.error('동행복권 API 호출 실패 및 로컬 데이터 없음: %s', e, exc_info=True)
        return None, '동행복권 API에서 당첨 정보를 가져오지 못했습니다. 잠시 후 다시 시도해주세요.'


# --- 라우트 ---
@app.route('/favicon.ico')
def favicon():
    return '', 204, {'Content-Type': 'image/x-icon'}


def _get_lotto645_data_row_count():
    """Lotto645.xlsx 데이터 행 수(헤더 제외). openpyxl 없으면 None."""
    try:
        import openpyxl
        p = (BASE_DIR / '.source' / 'Lotto645.xlsx').resolve()
        if not p.is_file():
            return None
        wb = openpyxl.load_workbook(p, read_only=True)
        ws = wb.active
        n = max(0, ws.max_row - 1) if ws.max_row else 0
        wb.close()
        return n
    except Exception:
        return None


def _fetch_rounds_draw_info(round_list):
    """
    누락 회차 목록을 독립형 API(get_lotto_round.get_lotto_number)로 조회해 추첨정보만 반환.
    Excel 미사용. 반환: [ { drwNo, date, numbers, bonus }, ... ]
    """
    try:
        from utils.get_lotto_round import get_lotto_number
    except ImportError as e:
        logger.error('[로또] utils.get_lotto_round 임포트 실패: %s', e)
        return []
    out = []
    for rno in round_list:
        try:
            rno = int(rno)
        except (TypeError, ValueError):
            continue
        try:
            result = get_lotto_number(rno)
        except Exception as e:
            logger.error('[로또] 회차 %s 조회 예외: %s', rno, e)
            continue
        if result is not None:
            out.append({
                'drwNo': result.get('drwNo'),
                'date': result.get('date') or '',
                'numbers': result.get('numbers') or [],
                'bonus': result.get('bonus'),
            })
        else:
            logger.warning('[로또] 회차 %s: 동행복권 API에서 데이터 없음(또는 접속 실패)', rno)
    if round_list and not out:
        logger.warning('[로또] 누락 회차 %s건 조회 결과 0건. 동행복권 접속/방화벽 확인 필요.', len(round_list))
    return out


@app.route('/api/health', methods=['GET', 'OPTIONS'])
def api_health():
    """Flask 서버 연결 확인용. 404가 나오면 server.py가 아닌 다른 서버가 8000번에서 동작 중."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS
    return jsonify(returnValue='ok', server='Flask (server.py)', startTime=SERVER_START_TIME), 200, CORS_HEADERS


@app.route('/api/fetch-missing-rounds', methods=['GET', 'OPTIONS'])
def api_fetch_missing_rounds():
    """누락 회차 추첨정보만 조회(Excel 미수정). 쿼리: rounds=1201,1202,1209"""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS
    rounds_param = request.args.get('rounds', '')
    round_list = [x.strip() for x in rounds_param.split(',') if x.strip()]
    fetched = _fetch_rounds_draw_info(round_list)
    return jsonify(returnValue='success', rounds=fetched), 200, CORS_HEADERS


def _sync_lotto645_xlsx():
    """
    Lotto645.xlsx 검증 후 빠진 회차(1~동행복권 최신회차)를 API로 조회해 추가.
    반환: (success, added, total_rounds, error_msg)
    """
    try:
        import openpyxl
        from utils.get_lotto_round import get_latest_round_no, get_lotto_number
    except ImportError as e:
        return False, 0, 0, 'openpyxl 또는 utils.get_lotto_round 미설치: %s' % e, []

    xlsx_path = (BASE_DIR / '.source' / 'Lotto645.xlsx').resolve()
    if not xlsx_path.is_file():
        return False, 0, 0, 'Lotto645.xlsx 파일이 없습니다.', []

    latest = get_latest_round_no()
    if latest is None:
        return False, 0, 0, '동행복권 API에서 최신 회차를 가져오지 못했습니다.', []

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return False, 0, 0, 'Lotto645.xlsx 열기 실패(Excel에서 닫아 주세요): %s' % str(e)[:80], []

    ws = wb.active
    if ws.max_row < 2:
        wb.close()
        return False, 0, 0, 'Lotto645.xlsx에 헤더만 있거나 비어 있습니다.', []

    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_row = [str(h).strip() if h is not None else '' for h in header_row]
    idx_round = idx_date = idx_bonus = None
    idx_nums = []
    for i, h in enumerate(header_row):
        c = i + 1
        if h == '회차':
            idx_round = c
        elif h == '날짜':
            idx_date = c
        elif h == '보너스번호':
            idx_bonus = c
        elif h and len(h) == 3 and h.startswith('번호') and h[2].isdigit():
            idx_nums.append((int(h[2]), c))
    idx_nums.sort(key=lambda x: x[0])
    idx_nums = [col for _, col in idx_nums]
    if idx_round is None or idx_date is None or len(idx_nums) != 6 or idx_bonus is None:
        wb.close()
        return False, 0, 0, 'Lotto645.xlsx 헤더(회차, 날짜, 번호1~6, 보너스번호)를 찾을 수 없습니다.', []

    existing_rounds = set()
    rows_by_round = {}
    for r in range(2, ws.max_row + 1):
        try:
            rno = int(ws.cell(r, idx_round).value)
        except (TypeError, ValueError):
            continue
        existing_rounds.add(rno)
        row_vals = [ws.cell(r, col).value for col in range(1, ws.max_column + 1)]
        rows_by_round[rno] = row_vals

    need_rounds = sorted(set(range(1, latest + 1)) - existing_rounds)
    if not need_rounds:
        wb.close()
        return True, 0, latest, None, []

    added = 0
    added_rounds = []
    for rno in need_rounds:
        result = get_lotto_number(rno)
        if result is None:
            continue
        row_vals = [None] * ws.max_column
        row_vals[idx_round - 1] = result.get('drwNo')
        row_vals[idx_date - 1] = result.get('date') or ''
        nums = result.get('numbers', [])
        for i, col in enumerate(idx_nums):
            row_vals[col - 1] = nums[i] if i < len(nums) else None
        row_vals[idx_bonus - 1] = result.get('bonus')
        rows_by_round[rno] = row_vals
        added += 1
        added_rounds.append({
            'drwNo': result.get('drwNo'),
            'date': result.get('date') or '',
            'numbers': nums,
            'bonus': result.get('bonus'),
        })

    # 회차 내림차순(최신 회차 먼저)으로 저장
    sorted_rounds = sorted(rows_by_round.keys(), reverse=True)
    while ws.max_row >= 2:
        ws.delete_rows(2, 1)
    for row_idx, rno in enumerate(sorted_rounds, start=2):
        row_vals = rows_by_round[rno]
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(row_idx, col_idx, value=val)

    try:
        wb.save(xlsx_path)
        wb.close()
    except PermissionError:
        wb.close()
        return False, 0, 0, 'Lotto645.xlsx 저장 실패(파일 잠김). Excel에서 닫고 다시 시도하세요.', []
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return False, 0, 0, '저장 실패: %s' % str(e)[:80], []

    # JSON도 갱신
    try:
        from utils.convert_to_json import convert_xlsx_to_json
        convert_xlsx_to_json()
    except Exception as e:
        logger.error('[Lotto645] JSON 변환 실패: %s', e)

    return True, added, len(sorted_rounds), None, added_rounds


# Lotto645.xlsx 확장 컬럼 (추첨정보: 당첨금·당첨자수·총액)
LOTTO645_EXTRA_HEADERS = ['1등 당첨금액', '1등 당첨자 수', '1등 총당첨금액', '전체 당첨상금 총액']


def _update_latest_lotto645(pre_fetched_data=None):
    """
    최신회차 추첨정보를 Lotto645.xlsx 해당 행에 업데이트합니다.
    pre_fetched_data: 이미 API로 가져온 데이터 딕셔너리 (없으면 내부에서 조회)
    반환: (success, error_msg)
    """
    try:
        import openpyxl
    except ImportError:
        return False, 'openpyxl이 필요합니다.'

    if pre_fetched_data:
        data = pre_fetched_data
        # _to_latest_response 포맷일 수 있으므로 키 매핑 확인 필요
        # 하지만 여기선 _fetch_dhlottery_api 결과(parsed)를 그대로 받는다고 가정
    else:
        # 데이터가 없으면 직접 조회
        data, err = fetch_latest_lotto()
        if err or not data:
            return False, (err or '최신 추첨정보를 가져오지 못했습니다.')
        # fetch_latest_lotto는 _to_latest_response로 변환된 값을 리턴함.
        # 내부 로직엔 원본 키(rnk1WnAmt 등)가 필요할 수 있으나,
        # _to_latest_response 결과물에는 firstWinamnt 등이 있으므로 이를 활용.

    # 데이터 유효성 체크
    drw_no = data.get('drwNo')
    if not drw_no:
        return False, '회차 정보가 없습니다.'

    xlsx_path = (BASE_DIR / '.source' / 'Lotto645.xlsx').resolve()
    if not xlsx_path.is_file():
        return False, 'Lotto645.xlsx 파일이 없습니다.'

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        return False, 'Lotto645.xlsx 열기 실패(Excel에서 닫아 주세요): %s' % str(e)[:80]

    ws = wb.active
    if ws.max_row < 1:
        wb.close()
        return False, 'Lotto645.xlsx에 헤더가 없습니다.'

    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_row = [str(h).strip() if h is not None else '' for h in header_row]

    idx_round = idx_date = idx_bonus = None
    idx_nums = []
    for i, h in enumerate(header_row):
        c = i + 1
        if h == '회차':
            idx_round = c
        elif h == '날짜':
            idx_date = c
        elif h == '보너스번호':
            idx_bonus = c
        elif h and len(h) == 3 and h.startswith('번호') and h[2].isdigit():
            idx_nums.append((int(h[2]), c))
    idx_nums.sort(key=lambda x: x[0])
    idx_nums = [col for _, col in idx_nums]
    if idx_round is None or idx_date is None or len(idx_nums) != 6 or idx_bonus is None:
        wb.close()
        return False, 'Lotto645.xlsx 헤더(회차, 날짜, 번호1~6, 보너스번호)를 찾을 수 없습니다.'

    # 확장 헤더(1등 당첨금액 등) 없으면 추가
    need_extra = LOTTO645_EXTRA_HEADERS
    extra_indices = []
    for eh in need_extra:
        try:
            ci = header_row.index(eh) + 1
        except ValueError:
            ci = None
        extra_indices.append(ci)
    if any(i is None for i in extra_indices):
        for eh in need_extra:
            if eh not in header_row:
                header_row.append(eh)
        max_col = len(header_row)
        for c in range(1, max_col + 1):
            ws.cell(1, c, value=header_row[c - 1] if c <= len(header_row) else None)
        for r in range(2, ws.max_row + 1):
            for c in range(ws.max_column + 1, max_col + 1):
                ws.cell(r, c, value=None)
        idx_1amt = header_row.index('1등 당첨금액') + 1
        idx_1cnt = header_row.index('1등 당첨자 수') + 1
        idx_1sum = header_row.index('1등 총당첨금액') + 1
        idx_tot = header_row.index('전체 당첨상금 총액') + 1
    else:
        idx_1amt, idx_1cnt, idx_1sum, idx_tot = extra_indices

    num_cols = max(ws.max_column, len(header_row))
    
    # 데이터 매핑 (pre_fetched_data가 parsed 원본인지, _to_latest_response 변환본인지에 따라 다름)
    # 여기서는 fetch_latest_lotto() 내부에서 호출 시 원본 parsed를 넘기도록 수정할 예정이므로
    # 원본 키와 변환 키를 모두 고려
    
    row_vals = [None] * num_cols
    row_vals[idx_round - 1] = drw_no
    row_vals[idx_date - 1] = data.get('drwNoDate') or ''
    
    # 당첨번호 매핑
    # parsed 원본: main=[1,2,3,4,5,6] 형태 / _to_latest_response: drwtNo1...drwtNo6
    if 'main' in data:
        nums = data['main']
    else:
        nums = [data.get(f'drwtNo{i}') for i in range(1, 7)]

    for i, col in enumerate(idx_nums):
        row_vals[col - 1] = nums[i] if i < len(nums) else None
        
    row_vals[idx_bonus - 1] = data.get('bnusNo')
    
    # 당첨금 매핑
    # 원본: rnk1WnAmt 등, 변환본: firstWinamnt 등
    # 단, 엑셀에는 '숫자'로 넣어야 함. 포맷팅된 문자열(x,xxx원)이 아님.
    
    def get_int(k1, k2):
        v = data.get(k1)
        if v is None: v = data.get(k2)
        try:
            return int(v)
        except (TypeError, ValueError):
            return None
            
    row_vals[idx_1amt - 1] = get_int('firstWinamnt', 'rnk1WnAmt')
    row_vals[idx_1cnt - 1] = get_int('firstPrzwnerCo', 'rnk1WnNope')
    row_vals[idx_1sum - 1] = get_int('firstAccumamnt', 'rnk1SumWnAmt')
    row_vals[idx_tot - 1] = get_int('totSellamnt', 'wholEpsdSumNtslAmt')

    target_row = None
    for r in range(2, ws.max_row + 1):
        try:
            if int(ws.cell(r, idx_round).value) == int(drw_no):
                target_row = r
                break
        except (TypeError, ValueError):
            continue

    if target_row is not None:
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(target_row, col_idx, value=val)
    else:
        # 최신 데이터를 맨 위(2행)에 삽입
        ws.insert_rows(2, 1)
        for col_idx, val in enumerate(row_vals, start=1):
            ws.cell(2, col_idx, value=val)

    try:
        wb.save(xlsx_path)
        wb.close()
    except PermissionError:
        wb.close()
        return False, 'Lotto645.xlsx 저장 실패(파일 잠김). Excel에서 닫고 다시 시도하세요.'
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return False, '저장 실패: %s' % str(e)[:80]

    # JSON도 갱신
    try:
        from utils.convert_to_json import convert_xlsx_to_json
        convert_xlsx_to_json()
    except Exception as e:
        logger.error('[Lotto645] JSON 변환 실패: %s', e)

    logger.info('[Lotto645] 최신회차 %s회 추첨정보로 업데이트 완료.', drw_no)
    return True, None


@app.route('/api/sync-lotto645', methods=['GET', 'POST', 'OPTIONS'])
def api_sync_lotto645():
    """Lotto645.xlsx에 빠진 회차(1~동행복권 최신) 추가. Excel 마지막회차 < API 최신회차일 때 호출."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS
    result = _sync_lotto645_xlsx()
    success, added, total_rounds, err, added_rounds = (result[0], result[1], result[2], result[3], result[4] if len(result) > 4 else [])
    if err:
        return jsonify(returnValue='fail', error=err, added=0, totalRounds=0, addedRounds=[]), 200, CORS_HEADERS
    return jsonify(returnValue='success', added=added, totalRounds=total_rounds, addedRounds=added_rounds), 200, CORS_HEADERS


@app.route('/api/lotto645-meta', methods=['GET', 'OPTIONS'])
def api_lotto645_meta():
    """Lotto645.xlsx 현재 데이터 행 수 반환. 클라이언트가 캐시 여부 검증용으로 사용."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS
    n = _get_lotto645_data_row_count()
    out = {'dataRows': n} if n is not None else {}
    headers = {**CORS_HEADERS, 'Cache-Control': 'no-store, no-cache, max-age=0', 'Pragma': 'no-cache'}
    return jsonify(out), 200, headers


@app.route('/api/lotto-latest', methods=['GET', 'OPTIONS'])
def api_lotto_latest():
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS
    force = request.args.get('force', '').lower() in ('1', 'true', 'yes')
    data, err = fetch_latest_lotto(force=force)
    if err:
        return jsonify(returnValue='fail', error=err), 200, CORS_HEADERS
    return jsonify(data), 200, CORS_HEADERS


def _get_round_from_lotto645_xlsx(round_no):
    """
    Lotto645.xlsx에서 특정 회차 행을 읽어 API 응답과 동일한 형식(parsed)으로 반환.
    동행복권 API 실패 시 fallback으로 사용. 반환: dict 또는 None
    """
    try:
        import openpyxl
    except ImportError:
        return None
    xlsx_path = (BASE_DIR / '.source' / 'Lotto645.xlsx').resolve()
    if not xlsx_path.is_file():
        return None
    try:
        rno = int(round_no)
    except (TypeError, ValueError):
        return None

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    except Exception:
        return None

    ws = wb.active
    if ws.max_row < 2:
        wb.close()
        return None

    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_row = [str(h).strip() if h is not None else '' for h in header_row]

    idx_round = idx_date = idx_bonus = None
    idx_nums = []
    for i, h in enumerate(header_row):
        c = i + 1
        if h == '회차':
            idx_round = c
        elif h == '날짜':
            idx_date = c
        elif h == '보너스번호':
            idx_bonus = c
        elif h and len(h) == 3 and h.startswith('번호') and h[2].isdigit():
            idx_nums.append((int(h[2]), c))
    idx_nums.sort(key=lambda x: x[0])
    idx_nums = [col for _, col in idx_nums]
    if idx_round is None or idx_date is None or len(idx_nums) != 6 or idx_bonus is None:
        wb.close()
        return None

    idx_1amt = idx_1cnt = idx_1sum = idx_tot = None
    for eh in LOTTO645_EXTRA_HEADERS:
        try:
            idx_1amt = header_row.index('1등 당첨금액') + 1
            idx_1cnt = header_row.index('1등 당첨자 수') + 1
            idx_1sum = header_row.index('1등 총당첨금액') + 1
            idx_tot = header_row.index('전체 당첨상금 총액') + 1
            break
        except ValueError:
            pass

    import datetime
    for r in range(2, ws.max_row + 1):
        try:
            cell_val = ws.cell(r, idx_round).value
            if cell_val is None:
                continue
            if int(cell_val) != rno:
                continue
        except (TypeError, ValueError):
            continue
        
        date_val = ws.cell(r, idx_date).value
        if isinstance(date_val, (datetime.datetime, datetime.date)):
             date_str = date_val.strftime('%Y-%m-%d')
        else:
             date_str = str(date_val).strip() if date_val is not None else ''
        
        # 날짜 문자열에 시간 포함 시 제거 (예: 2025-01-01 00:00:00)
        if ' ' in date_str:
            date_str = date_str.split(' ')[0]

        main = []
        for col in idx_nums:
            v = ws.cell(r, col).value
            try:
                n = int(v) if v is not None and str(v).strip() != '' else None
                main.append(n)
            except (TypeError, ValueError):
                main.append(None)
        if None in main or len(main) != 6:
            continue
        bns = ws.cell(r, idx_bonus).value
        try:
            bnus_no = int(bns) if bns is not None and str(bns).strip() != '' else None
        except (TypeError, ValueError):
            bnus_no = None
        if bnus_no is None:
            continue

        first_win = first_cnt = first_sum = tot_sell = None
        if idx_1amt is not None:
            try:
                v = ws.cell(r, idx_1amt).value
                first_win = int(v) if v is not None and str(v).strip() != '' else None
            except (TypeError, ValueError):
                pass
        if idx_1cnt is not None:
            try:
                v = ws.cell(r, idx_1cnt).value
                first_cnt = int(v) if v is not None and str(v).strip() != '' else None
            except (TypeError, ValueError):
                pass
        if idx_1sum is not None:
            try:
                v = ws.cell(r, idx_1sum).value
                first_sum = int(v) if v is not None and str(v).strip() != '' else None
            except (TypeError, ValueError):
                pass
        if idx_tot is not None:
            try:
                v = ws.cell(r, idx_tot).value
                tot_sell = int(v) if v is not None and str(v).strip() != '' else None
            except (TypeError, ValueError):
                pass

        wb.close()
        return {
            'drwNo': rno,
            'drwNoDate': date_str,
            'main': main,
            'bnusNo': bnus_no,
            'rnk1WnAmt': first_win,  # _to_latest_response 호환
            'firstWinamnt': first_win,
            'firstPrzwnerCo': first_cnt,
            'firstAccumamnt': first_sum,
            'totSellamnt': tot_sell,
            'source': 'Lotto645.xlsx (로컬)',
        }

    wb.close()
    return None


# --- Google Gemini AI 설정 ---
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
if not GEMINI_API_KEY:
    logger.warning('[경고] GEMINI_API_KEY가 .env 파일에 설정되지 않았습니다.')


@app.route('/api/ask-gemini', methods=['POST', 'OPTIONS'])
def api_ask_gemini():
    """Lotto645 최근 10회 데이터를 바탕으로 AI에게 질문"""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS

    if not GEMINI_API_KEY:
        logger.error('[Gemini API] API 키가 설정되지 않았습니다.')
        return jsonify(returnValue='fail', error='API 키가 설정되지 않았습니다. .env 파일을 확인하세요.'), 200, CORS_HEADERS
        

    try:
        req_data = request.get_json() or {}
        user_question = req_data.get('question', '최근 로또 패턴 분석해줘')
        start_round_param = req_data.get('startRound')
        end_round_param = req_data.get('endRound')
        target_rounds = req_data.get('targetRounds')  # 회차별 당첨조회 기준 필터링된 목록
        
        # 최근 로또 데이터 로드 (JSON)
        json_path = (BASE_DIR / '.source' / 'Lotto645.json').resolve()
        
        recent_data = []
        analyze_count = 30  # 기본 분석 데이터 수 (최근 30회)

        if json_path.is_file():
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data:
                    # 1. 숫자형 회차 변환 및 정렬
                    for item in data:
                        item['_round_int'] = int(item.get('회차', 0))
                    
                    data.sort(key=lambda x: x['_round_int'], reverse=True)

                    max_analyze_count = 30
                    try:
                        # 1. 특정 회차 목록이 전달된 경우 (회차별 당첨조회 기준)
                        if target_rounds and isinstance(target_rounds, list):
                            round_set = set(int(r) for r in target_rounds)
                            # JSON 데이터에서 해당 회차들만 추출하여 최신순 정렬
                            recent_data = [item for item in data if item['_round_int'] in round_set]
                            # 분석 속도를 위해 최근 N개로 제한
                            if len(recent_data) > max_analyze_count:
                                recent_data = recent_data[:max_analyze_count]
                        
                        # 2. 시작/종료 범위가 지정된 경우
                        elif start_round_param and end_round_param:
                            s_target = int(start_round_param)
                            e_target = int(end_round_param)
                            recent_data = [item for item in data if s_target <= item['_round_int'] <= e_target]
                            if len(recent_data) > max_analyze_count:
                                recent_data = recent_data[:max_analyze_count]
                        
                        # 3. 기본값 (최신 N회차)
                        else:
                            recent_data = data[:analyze_count]
                    except (ValueError, TypeError):
                        recent_data = data[:analyze_count]

                    # 3. 시계열 분석을 위해 과거 -> 최신 순(오름차순)으로 재정렬
                    recent_data.sort(key=lambda x: x['_round_int'])
        
        if not recent_data:
            return jsonify(returnValue='fail', error='분석할 로또 데이터가 없습니다.'), 200, CORS_HEADERS

        analysis_range_text = f"{recent_data[0]['_round_int']}회 ~ {recent_data[-1]['_round_int']}회" if len(recent_data) > 1 else f"{recent_data[0]['_round_int']}회"
        history_text = f"분석 대상 범위: {analysis_range_text} (총 {len(recent_data)}회차)\n"
        history_text += "로또 당첨번호 이력 (과거 -> 최신순):\n"
        for row in recent_data:
            nums = [row.get(f'번호{i}') for i in range(1, 7)]
            history_text += f"- {row.get('회차')}회 ({row.get('날짜')}): {nums} + 보너스 {row.get('보너스번호')}\n"
        
        # AI 분석 모듈 호출
        try:
            from utils.ai_analysis import analyze_lotto_patterns
            answer = analyze_lotto_patterns(history_text, user_question, GEMINI_API_KEY)
            
            if answer:
                return jsonify(returnValue='success', answer=answer), 200, CORS_HEADERS
            else:
                return jsonify(returnValue='fail', error='AI 응답이 비어있습니다.'), 200, CORS_HEADERS
                
        except ImportError as e:
             logger.error(f'[Server] AI 모듈 로드 실패: {e}')
             return jsonify(returnValue='fail', error='AI 분석 모듈을 로드할 수 없습니다.'), 500, CORS_HEADERS
        except Exception as e:
             # AI 호출 실패 시 로그 남기고 클라이언트엔 간단히 전달
             logger.error(f'[Gemini API 오류] %s', e, exc_info=True)
             return jsonify(returnValue='fail', error=f'AI 분석 중 오류가 발생했습니다: {str(e)[:100]}'), 200, CORS_HEADERS

    except Exception as e:
        logger.error(f'[API /api/ask-gemini 오류] %s', e, exc_info=True)
        return jsonify(returnValue='fail', error=f'서버 내부 오류: {str(e)[:100]}'), 200, CORS_HEADERS


@app.route('/api/lotto-round/<int:round_no>', methods=['GET', 'OPTIONS'])
def api_lotto_round(round_no):
    """특정 회차 추첨정보 (동행복권 최신 추첨정보와 동일 형식). API 실패 시 Lotto645.xlsx 로컬 fallback."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS
    
    from utils.get_lotto_round import get_lotto_number
    parsed = get_lotto_number(round_no)
    if not parsed:
        parsed = _get_round_from_lotto645_xlsx(round_no)
    if not parsed:
        return jsonify(returnValue='fail', error='해당 회차(%s) 추첨정보를 가져오지 못했습니다.' % round_no), 200, CORS_HEADERS
    return jsonify(_to_latest_response(parsed)), 200, CORS_HEADERS


@app.route('/api/save-lotto023', methods=['POST', 'OPTIONS'])
def api_save_lotto023():
    """Lotto023 게임 데이터를 Lotto023.xlsx에 저장 (세트/게임 자동 관리)."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS

    try:
        data = request.get_json(silent=True)
        if not data or 'games' not in data:
            return jsonify(returnValue='fail', error='저장할 데이터가 없거나 형식이 올바르지 않습니다.'), 400, CORS_HEADERS

        import openpyxl
        source_dir = BASE_DIR / '.source'
        source_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = (source_dir / 'Lotto023.xlsx').resolve()
        
        default_headers = list(LOTTO023_CANONICAL_HEADERS)

        # 파일이 없으면 새로 생성
        if not xlsx_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Lotto023"
            for c, header in enumerate(default_headers, 1):
                ws.cell(1, c, value=header)
            wb.save(xlsx_path)
            wb.close()

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # 기존 데이터 모두 로드
        rows = list(ws.iter_rows(values_only=True))
        data_rows = []
        header_in_file = []
        if rows:
            header_in_file = [str(h) if h is not None else '' for h in rows[0]]
            if len(rows) > 1:
                data_rows = [list(r) for r in rows[1:]]
        
        # '세트' 컬럼 인덱스 확인 (없으면 추가해야 함)
        if '세트' not in header_in_file:
            header_in_file.insert(1, '세트')
            for r in data_rows:
                r.insert(1, '1')

        # 표준 헤더로 통일 (조회시작/종료 제거, 선택합계 보강)
        header_in_file, data_rows = _migrate_lotto023_rows(header_in_file, data_rows)

        headers = header_in_file

        idx_round = headers.index('회차')
        idx_set = headers.index('세트')
        idx_game = headers.index('게임')

        new_games = data['games']
        added_count = 0
        
        # 각 게임별로 세트/게임 번호 부여
        for g in new_games:
            if not _lotto023_pick_valid(g):
                logger.warning('[Lotto023] 유효하지 않은 선택 번호로 저장 거절: %s', g)
                wb.close()
                return jsonify(
                    returnValue='fail',
                    error='유효하지 않은 선택 번호입니다. 선택1~6은 1~45의 서로 다른 숫자 6개여야 합니다.',
                ), 400, CORS_HEADERS
            raw_round = g.get('회차')
            if not raw_round:
                logger.warning('[Lotto023] 게임 데이터에 회차 정보가 누락되었습니다: %s', g)
                continue
            
            try:
                target_round = int(raw_round)
            except (ValueError, TypeError):
                logger.warning('[Lotto023] 유효하지 않은 회차 형식입니다: %s', raw_round)
                continue
            
            # 해당 회차의 기존 데이터 찾기
            round_data = []
            for r in data_rows:
                try:
                    if int(r[idx_round]) == target_round:
                        round_data.append(r)
                except: continue
            
            if not round_data:
                # 회차가 없으면 1세트, 1게임부터 시작
                next_set = 1
                next_game = 1
            else:
                # 회차가 있으면 마지막 세트/게임 확인
                # 세트/게임 순서대로 정렬하여 마지막 값 추출
                def sort_key(x):
                    try:
                        s = int(x[idx_set]) if x[idx_set] is not None else 1
                        g_num = int(x[idx_game]) if x[idx_game] is not None else 1
                        return (s, g_num)
                    except: return (0, 0)
                
                round_data.sort(key=sort_key)
                last_row = round_data[-1]
                
                try:
                    last_set = int(last_row[idx_set]) if last_row[idx_set] is not None else 1
                    last_game = int(last_row[idx_game]) if last_row[idx_game] is not None else 1
                    
                    if last_game >= 5:
                        next_set = last_set + 1
                        next_game = 1
                    else:
                        next_set = last_set
                        next_game = last_game + 1
                except:
                    next_set = 1
                    next_game = 1

            # 새 행 생성
            new_row = [None] * len(headers)
            for i, h in enumerate(headers):
                if h == '회차':
                    new_row[i] = str(target_round)
                elif h == '세트':
                    new_row[i] = str(next_set)
                elif h == '게임':
                    new_row[i] = str(next_game)
                elif h == '선택합계':
                    sv = g.get('선택합계')
                    if sv is not None and str(sv).strip() != '':
                        new_row[i] = str(sv).strip()
                    else:
                        try:
                            s = sum(int(g.get('선택%d' % j, 0) or 0) for j in range(1, 7))
                            new_row[i] = str(s) if s else ''
                        except (ValueError, TypeError):
                            new_row[i] = ''
                elif h == 'Perfect순위':
                    pv = g.get('Perfect순위')
                    new_row[i] = '' if pv is None or str(pv).strip() == '' else str(pv).strip()
                else:
                    val = g.get(h, '')
                    new_row[i] = '' if val is None else val
            
            data_rows.append(new_row)
            added_count += 1
            
        # 전체 데이터 정렬 (회차 내림차순, 세트 오름차순, 게임 오름차순)
        def final_sort_key(x):
            try:
                r = -int(x[idx_round]) if x[idx_round] else 0
                s = int(x[idx_set]) if x[idx_set] else 0
                g_num = int(x[idx_game]) if x[idx_game] else 0
                return (r, s, g_num)
            except: return (0, 0, 0)
            
        data_rows.sort(key=final_sort_key)
            
        # 시트 비우고 다시 쓰기
        ws.delete_rows(1, ws.max_row)
        for c, header in enumerate(headers, 1):
            ws.cell(1, c, value=header)
            
        for r_idx, r_data in enumerate(data_rows, 2):
            for c_idx, val in enumerate(r_data, 1):
                ws.cell(r_idx, c_idx, value=val)
                
        wb.save(xlsx_path)
        wb.close()

        return jsonify(returnValue='success', count=added_count), 200, CORS_HEADERS

    except PermissionError:
        return jsonify(returnValue='fail', error='Lotto023.xlsx 파일이 다른 프로그램에서 열려 있습니다.'), 200, CORS_HEADERS
    except Exception as e:
        logger.error('[Lotto023] 저장 중 오류: %s', e)
        return jsonify(returnValue='fail', error=str(e)), 200, CORS_HEADERS


@app.route('/api/rewrite-lotto023', methods=['POST', 'OPTIONS'])
def api_rewrite_lotto023():
    """Lotto023.xlsx 전체 데이터 행을 요청 본문의 games(표준 컬럼 dict 목록)로 교체."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS

    try:
        data = request.get_json(silent=True)
        if not data or 'games' not in data:
            return jsonify(returnValue='fail', error='games 배열이 필요합니다.'), 400, CORS_HEADERS

        import openpyxl
        games = data['games']
        if not isinstance(games, list):
            return jsonify(returnValue='fail', error='games는 배열이어야 합니다.'), 400, CORS_HEADERS

        source_dir = BASE_DIR / '.source'
        source_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = (source_dir / 'Lotto023.xlsx').resolve()
        canon = list(LOTTO023_CANONICAL_HEADERS)
        idx_round = canon.index('회차')
        idx_set = canon.index('세트')
        idx_game = canon.index('게임')

        data_rows = []
        for g in games:
            if not isinstance(g, dict):
                continue
            if not _lotto023_pick_valid(g):
                return jsonify(
                    returnValue='fail',
                    error='rewrite: 유효하지 않은 선택 번호가 있습니다. (회차 %s)' % g.get('회차', '?'),
                ), 400, CORS_HEADERS
            row = []
            for h in canon:
                v = g.get(h)
                if v is None or v == '':
                    row.append('')
                elif isinstance(v, str):
                    row.append(v.strip())
                else:
                    row.append(str(v))
            data_rows.append(row)

        def sort_key(x):
            try:
                r = -int(x[idx_round]) if x[idx_round] not in (None, '') else 0
                s = int(x[idx_set]) if x[idx_set] not in (None, '') else 0
                gn = int(x[idx_game]) if x[idx_game] not in (None, '') else 0
                return (r, s, gn)
            except (ValueError, TypeError):
                return (0, 0, 0)

        data_rows.sort(key=sort_key)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Lotto023'
        for c, header in enumerate(canon, 1):
            ws.cell(1, c, value=header)
        for r_idx, r_data in enumerate(data_rows, 2):
            for c_idx, val in enumerate(r_data, 1):
                ws.cell(r_idx, c_idx, value=val)
        wb.save(xlsx_path)
        wb.close()

        return jsonify(returnValue='success', count=len(data_rows)), 200, CORS_HEADERS

    except PermissionError:
        return jsonify(returnValue='fail', error='Lotto023.xlsx 파일이 다른 프로그램에서 열려 있습니다.'), 200, CORS_HEADERS
    except Exception as e:
        logger.error('[Lotto023] rewrite 오류: %s', e)
        return jsonify(returnValue='fail', error=str(e)), 200, CORS_HEADERS


def _perfect_json_row_dict_from_game(g, canon):
    if not isinstance(g, dict):
        return None
    row = {}
    for h in canon:
        v = g.get(h)
        if v is None or v == '':
            row[h] = ''
        elif isinstance(v, str):
            row[h] = v.strip()
        else:
            row[h] = str(v)
    return row


def _perfect_json_sort_key(d):
    try:
        r = -(int(d.get('회차') or 0))
        s = int(d.get('세트') or 0)
        gn = int(d.get('게임') or 0)
        return (r, s, gn)
    except (ValueError, TypeError):
        return (0, 0, 0)


@app.route('/api/save-perfect', methods=['POST', 'OPTIONS'])
def api_save_perfect():
    """B of B 후보를 .source/perfect.json에 요청 games로 전체 교체 (Lotto023과 동일 표준 키)."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS

    try:
        data = request.get_json(silent=True)
        if not data or 'games' not in data:
            return jsonify(returnValue='fail', error='games 배열이 필요합니다.'), 400, CORS_HEADERS

        games = data['games']
        if not isinstance(games, list):
            return jsonify(returnValue='fail', error='games는 배열이어야 합니다.'), 400, CORS_HEADERS

        source_dir = BASE_DIR / '.source'
        source_dir.mkdir(parents=True, exist_ok=True)
        json_path = (source_dir / 'perfect.json').resolve()
        canon = list(LOTTO023_CANONICAL_HEADERS)

        out_games = []
        for g in games:
            if not _lotto023_pick_valid(g):
                return jsonify(
                    returnValue='fail',
                    error='perfect.json: 유효하지 않은 선택 번호가 있습니다. (회차 %s)' % g.get('회차', '?'),
                ), 400, CORS_HEADERS
            row = _perfect_json_row_dict_from_game(g, canon)
            if row:
                out_games.append(row)

        out_games.sort(key=_perfect_json_sort_key)

        payload = {'games': out_games}
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

        return jsonify(
            returnValue='success',
            count=len(out_games),
            savedRelativePath='.source/perfect.json',
        ), 200, CORS_HEADERS

    except PermissionError:
        return jsonify(returnValue='fail', error='B of B 후보(perfect.json) 파일을 쓸 수 없습니다. 다른 프로그램에서 사용 중일 수 있습니다.'), 200, CORS_HEADERS
    except Exception as e:
        logger.error('[perfect.json] 저장 오류: %s', e)
        return jsonify(returnValue='fail', error=str(e)), 200, CORS_HEADERS


@app.route('/api/delete-lotto023', methods=['POST', 'OPTIONS'])
def api_delete_lotto023():
    """Lotto023 게임 데이터를 Lotto023.xlsx에서 삭제."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS

    try:
        data = request.get_json(silent=True)
        if not data or 'items' not in data:
            return jsonify(returnValue='fail', error='삭제할 데이터 정보가 없습니다.'), 400, CORS_HEADERS

        import openpyxl
        xlsx_path = (BASE_DIR / '.source' / 'Lotto023.xlsx').resolve()
        if not xlsx_path.exists():
            return jsonify(returnValue='fail', error='Lotto023.xlsx 파일이 없습니다.'), 404, CORS_HEADERS

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else '' for c in range(1, ws.max_column + 1)]
        idx_round = headers.index('회차') if '회차' in headers else None
        idx_set = headers.index('세트') if '세트' in headers else None
        idx_game = headers.index('게임') if '게임' in headers else None

        if idx_round is None or idx_game is None:
             wb.close()
             return jsonify(returnValue='fail', error='필수 컬럼(회차, 게임)을 찾을 수 없습니다.'), 500, CORS_HEADERS

        items_to_del = data['items'] # [ {round, set, game}, ... ]
        
        # 삭제 대상 필터링 (메모리상에서 처리 후 시트 갱신이 안전)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        new_rows = []
        deleted_count = 0
        
        for r in rows:
            if not r or r[0] is None: continue
            
            is_match = False
            round_val = str(r[idx_round])
            set_val = str(r[idx_set]) if idx_set is not None else ''
            game_val = str(r[idx_game])
            
            for item in items_to_del:
                if str(item['round']) == round_val and str(item['game']) == game_val:
                    # 세트가 있으면 세트까지 비교
                    if idx_set is not None:
                        if str(item.get('set', '')) == set_val:
                            is_match = True
                            break
                    else:
                        is_match = True
                        break
            
            if is_match:
                deleted_count += 1
            else:
                new_rows.append(r)
        
        # 데이터 다시 쓰기
        ws.delete_rows(2, ws.max_row)
        for r_idx, r_data in enumerate(new_rows, 2):
            for c_idx, val in enumerate(r_data, 1):
                ws.cell(r_idx, c_idx, value=val)
        
        wb.save(xlsx_path)
        wb.close()

        return jsonify(returnValue='success', deleted=deleted_count), 200, CORS_HEADERS

    except Exception as e:
        logger.error('[Lotto023] 삭제 중 오류: %s', e)
        return jsonify(returnValue='fail', error=str(e)), 200, CORS_HEADERS


@app.route('/api/delete-all-lotto023', methods=['POST', 'OPTIONS'])
def api_delete_all_lotto023():
    """Lotto023 모든 데이터를 삭제."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS

    try:
        import os, openpyxl
        xlsx_path = os.path.join(FILE_DIR, 'Lotto023.xlsx')
        if not os.path.exists(xlsx_path):
            return jsonify(returnValue='success', message='파일이 없습니다.'), 200, CORS_HEADERS

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # 헤더(1행) 제외하고 내용이 있는 경우 삭제
        if ws.max_row > 1:
            # 모든 데이터 행 삭제 (2행부터 끝까지)
            ws.delete_rows(2, ws.max_row)
            wb.save(xlsx_path)
            wb.close()
            
        return jsonify(returnValue='success'), 200, CORS_HEADERS
    except Exception as e:
        logger.error('[Lotto023] 전체 삭제 중 오류: %s', e)
        return jsonify(returnValue='fail', error=str(e)), 200, CORS_HEADERS

@app.route('/api/save-ticket-desktop', methods=['POST', 'OPTIONS'])
def api_save_ticket_desktop():
    """프리미엄 티켓 이미지를 바탕화면에 저장."""
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS

    try:
        data = request.get_json(silent=True)
        if not data or 'image' not in data or 'round' not in data:
            return jsonify(returnValue='fail', error='데이터가 올바르지 않습니다.'), 400, CORS_HEADERS

        import base64
        import os
        from pathlib import Path

        desktop_path = Path.home() / 'Desktop'
        
        image_data = data['image']
        round_no = data['round']

        if ',' in image_data:
            image_data = image_data.split(',')[1]

        file_name = f"Lotto_Premium_Ticket_{round_no}회.png"
        file_path = desktop_path / file_name

        with open(file_path, "wb") as fh:
            fh.write(base64.b64decode(image_data))

        return jsonify(returnValue='success', path=str(file_path)), 200, CORS_HEADERS
    except Exception as e:
        logger.error('[save-ticket-desktop] 바탕화면 저장 오류: %s', e)
        return jsonify(returnValue='fail', error=str(e)), 500, CORS_HEADERS


# --- 정적 파일 (루트 및 하위 경로) ---
def _mimetype_with_charset(path):
    """한글 깨짐 방지: 텍스트 파일에 charset=utf-8 적용"""
    p = path.lower()
    if p.endswith('.html') or p.endswith('.htm'):
        return 'text/html; charset=utf-8'
    if p.endswith('.js'):
        return 'application/javascript; charset=utf-8'
    if p.endswith('.css'):
        return 'text/css; charset=utf-8'
    if p.endswith('.json'):
        return 'application/json; charset=utf-8'
    return None


@app.route('/api/shutdown', methods=['POST', 'OPTIONS'])
def api_shutdown():
    if request.method == 'OPTIONS':
        return '', 204, CORS_OPTIONS_HEADERS
    logger.info('[종료] 서버 종료 요청 수신')
    resp = jsonify({'returnValue': 'success', 'message': '서버를 종료합니다.'})
    resp.headers.update(CORS_HEADERS)

    import threading
    def _shutdown():
        import time
        time.sleep(0.5)
        os._exit(0)
    threading.Thread(target=_shutdown, daemon=True).start()
    return resp


@app.route('/')
def index():
    resp = send_from_directory(BASE_DIR, 'index.html')
    resp.headers['Content-Type'] = 'text/html; charset=utf-8'
    return resp


@app.route('/<path:path>')
def static_file(path):
    full = (BASE_DIR / path).resolve()
    base_resolved = BASE_DIR.resolve()
    if not str(full).startswith(str(base_resolved)) or not full.is_file():
        logger.warning('[404] 파일 없음: /%s (기대 경로: %s)', path, full)
        return '404 - 파일을 찾을 수 없습니다.', 404, {'Content-Type': 'text/plain; charset=utf-8'}
    resp = send_from_directory(BASE_DIR, path)
    mimetype = _mimetype_with_charset(path)
    if mimetype:
        resp.headers['Content-Type'] = mimetype
    # Lotto645.xlsx 등 데이터 파일은 매번 서버에서 새로 읽도록 캐시 비활성화 (브라우저/프록시 캐시 방지)
    path_lower = path.replace('\\', '/').lower()
    if path_lower.endswith('.xlsx') or 'lotto645.xlsx' in path_lower or 'lotto023.xlsx' in path_lower:
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
    return resp


if __name__ == '__main__':
    # 서버 시작 시 Lotto645 JSON 갱신
    try:
        from utils.convert_to_json import convert_xlsx_to_json
        logger.info('[초기화] Lotto645 XLSX -> JSON 변환 시작...')
        convert_xlsx_to_json()
    except Exception as e:
        logger.error('[초기화] Lotto645 JSON 변환 실패: %s', e)

    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() in ('1', 'true', 'yes')
    logger.info('서버 시작: http://localhost:%s/ (Debug: %s)', PORT, debug_mode)
    app.run(host='0.0.0.0', port=PORT, debug=debug_mode)
