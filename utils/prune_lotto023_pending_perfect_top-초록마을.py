# -*- coding: utf-8 -*-
"""
미추첨(당첨 데이터가 아직 없는) 회차 중, Lotto023에만 존재하는 가장 최신 회차 그룹에 대해
app.js의 Perfect 정렬(미추첨: 핫 22개 일치 수, 평균합 거리)과 동일한 기준으로 정렬한 뒤
상위 KEEP_N(기본 10)만 남기고 나머지 행을 Lotto023.xlsx에서 제거합니다.

※ 행운 모드는 브라우저의 조회기간(currentStatsRounds)을 알 수 없어,
  이 스크립트에서는 JS의 폴백과 같이 '해당 회차 미만 전체 회차'로 핫을 계산합니다.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

BASE = Path(__file__).resolve().parents[1]
SOURCE = BASE / ".source"
PATH_645 = SOURCE / "Lotto645.json"
PATH_023 = SOURCE / "Lotto023.xlsx"


def load_645_rounds(path: Path) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)
    out = []
    for item in raw:
        rno = int(item["회차"])
        nums = [int(item[f"번호{i}"]) for i in range(1, 7)]
        b = item.get("보너스번호")
        bonus = int(b) if b is not None and b != "" else None
        out.append({"round": rno, "numbers": nums, "bonus": bonus})
    out.sort(key=lambda x: -x["round"])
    return out


def calculate_win_stats(rounds: list[dict]) -> dict[int, int]:
    m = {i: 0 for i in range(1, 46)}
    for r in rounds:
        for n in r["numbers"]:
            if 1 <= n <= 45:
                m[n] = m.get(n, 0) + 1
    return m


def calculate_appearance_stats(rounds: list[dict]) -> dict[int, int]:
    m = {i: 0 for i in range(1, 46)}
    for r in rounds:
        for n in r["numbers"]:
            if 1 <= n <= 45:
                m[n] = m.get(n, 0) + 1
        b = r.get("bonus")
        if b is not None and 1 <= b <= 45:
            m[b] = m.get(b, 0) + 1
    return m


def calculate_consecutive_stats(rounds: list[dict]) -> dict[int, int]:
    m = {i: 0 for i in range(1, 46)}
    for r in rounds:
        nums = sorted(r["numbers"])
        for j in range(len(nums) - 1):
            if nums[j + 1] == nums[j] + 1:
                m[nums[j]] = m.get(nums[j], 0) + 1
                m[nums[j + 1]] = m.get(nums[j + 1], 0) + 1
    return m


def sort_and_split_hot_cold(
    win_map: dict[int, int],
    app_map: dict[int, int],
    seq_map: dict[int, int],
) -> tuple[list[int], list[int]]:
    rows = []
    for num in range(1, 46):
        rows.append(
            {
                "number": num,
                "count": win_map.get(num, 0),
                "appCount": app_map.get(num, 0),
                "seqCount": seq_map.get(num, 0),
            }
        )
    rows.sort(
        key=lambda s: (-s["count"], -s["appCount"], -s["seqCount"], -s["number"])
    )
    hot = [s["number"] for s in rows[:22]]
    cold = [s["number"] for s in rows[22:]][::-1]
    return hot, cold


def mean_six_sum(rounds: list[dict]) -> float:
    sums = []
    for r in rounds:
        n = r.get("numbers") or []
        if len(n) >= 6:
            sums.append(sum(int(x) for x in n[:6]))
    if not sums:
        return 138.0
    return sum(sums) / len(sums)


def stat_rounds_before(all_645: list[dict], round_num: int, mode: str) -> list[dict]:
    """JS 미추첨: AI 등은 전회차 미만. 행운은 조회기간 우선이나 스크립트는 전체 미만으로 통일(폴백과 동일)."""
    return [
        r
        for r in all_645
        if r["round"] < round_num and len(r.get("numbers") or []) >= 6
    ]


def perfect_key_pending(game_nums: list[int], round_num: int, mode: str, all_645: list[dict]):
    nums = sorted(int(x) for x in game_nums)
    valid = (
        len(nums) == 6
        and len(set(nums)) == 6
        and all(1 <= n <= 45 for n in nums)
    )
    game_sum = sum(nums) if valid else 99999

    stat_r = stat_rounds_before(all_645, round_num, mode)
    win_map = calculate_win_stats(stat_r)
    app_map = calculate_appearance_stats(stat_r)
    seq_map = calculate_consecutive_stats(stat_r)
    hot, _ = sort_and_split_hot_cold(win_map, app_map, seq_map)
    hot_set = set(hot)
    match_score = sum(1 for n in nums if n in hot_set) if valid else 0
    ref_avg = mean_six_sum(stat_r)
    sum_dist = abs(game_sum - ref_avg) if valid else 999999
    return match_score, sum_dist


def row_to_game(row: tuple, headers: list[str]) -> dict | None:
    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
    try:
        rno = int(d.get("회차") or 0)
        s = int(d.get("세트") or 1)
        g = int(d.get("게임") or 0)
    except (TypeError, ValueError):
        return None
    nums = []
    for i in range(1, 7):
        v = d.get(f"선택{i}")
        if v is None or v == "":
            return None
        try:
            nums.append(int(v))
        except (TypeError, ValueError):
            return None
    mode = (d.get("게임선택") or "기타").strip() or "기타"
    return {"round": rno, "set": s, "game": g, "numbers": nums, "mode": mode}


def make_cmp_perfect(all_645: list[dict], round_num: int):
    def cmp_games(a: dict, b: dict) -> int:
        ma, da = perfect_key_pending(a["numbers"], round_num, a["mode"], all_645)
        mb, db = perfect_key_pending(b["numbers"], round_num, b["mode"], all_645)
        if ma != mb:
            return -1 if ma > mb else (1 if ma < mb else 0)
        if da != db:
            return -1 if da < db else (1 if da > db else 0)
        if a["set"] != b["set"]:
            return -1 if a["set"] > b["set"] else (1 if a["set"] < b["set"] else 0)
        return -1 if a["game"] < b["game"] else (1 if a["game"] > b["game"] else 0)

    return cmp_games


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--keep", type=int, default=10, help="Perfect 상위 N개만 유지")
    ap.add_argument("--round", type=int, default=0, help="특정 회차만(0이면 자동: 최신 미추첨)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not PATH_645.exists():
        print("Lotto645.json 없음:", PATH_645)
        sys.exit(1)
    if not PATH_023.exists():
        print("Lotto023.xlsx 없음:", PATH_023)
        sys.exit(1)

    all_645 = load_645_rounds(PATH_645)
    latest_drawn = max(r["round"] for r in all_645)

    wb = openpyxl.load_workbook(PATH_023)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("빈 시트")
        sys.exit(0)
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows = [list(r) for r in rows[1:] if r and any(x is not None and x != "" for x in r)]

    idx_round = headers.index("회차")
    games_by_round: dict[int, list[tuple[list, dict]]] = {}
    for r in data_rows:
        g = row_to_game(tuple(r), headers)
        if not g:
            continue
        if g["round"] <= latest_drawn:
            continue
        games_by_round.setdefault(g["round"], []).append((r, g))

    if not games_by_round:
        print(f"미추첨(회차 > {latest_drawn}) 저장 게임이 없습니다.")
        wb.close()
        return

    if args.round:
        target = args.round
        if target not in games_by_round:
            print(f"회차 {target}에 미추첨 게임이 없습니다.")
            wb.close()
            sys.exit(1)
    else:
        target = max(games_by_round.keys())

    bucket = games_by_round[target]
    parsed = [x[1] for x in bucket]
    from functools import cmp_to_key

    parsed.sort(key=cmp_to_key(make_cmp_perfect(all_645, target)))

    keep_set = set()
    for item in parsed[: args.keep]:
        keep_set.add((item["round"], item["set"], item["game"]))

    to_remove = [x[0] for x in bucket if (x[1]["round"], x[1]["set"], x[1]["game"]) not in keep_set]

    print(
        f"Lotto645 최종 추첨 회차: {latest_drawn} → 미추첨 정리 대상 회차: {target} "
        f"(게임 {len(bucket)}개 중 상위 {args.keep}개 유지, 삭제 {len(to_remove)}개)"
    )

    if args.dry_run:
        for row in to_remove:
            print("  [dry-run] 삭제 예정 행:", row[idx_round], row[headers.index("세트")], row[headers.index("게임")])
        wb.close()
        return

    idx_set = headers.index("세트")
    idx_game = headers.index("게임")
    remove_keys = set()
    for r in to_remove:
        remove_keys.add((int(r[idx_round]), int(r[idx_set]), int(r[idx_game])))

    new_data = []
    for r in data_rows:
        try:
            rv = int(r[idx_round])
            sv = int(r[idx_set])
            gv = int(r[idx_game])
        except (TypeError, ValueError):
            new_data.append(r)
            continue
        if rv == target and (rv, sv, gv) in remove_keys:
            continue
        new_data.append(r)

    ws.delete_rows(1, ws.max_row)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, value=h)
    for r_idx, r_data in enumerate(new_data, 2):
        for c_idx, val in enumerate(r_data, 1):
            ws.cell(r_idx, c_idx, value=val)
    wb.save(PATH_023)
    wb.close()
    print("저장 완료:", PATH_023)


if __name__ == "__main__":
    main()
